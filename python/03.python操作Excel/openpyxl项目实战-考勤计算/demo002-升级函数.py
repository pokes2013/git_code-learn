import openpyxl
from openpyxl import load_workbook
from datetime import datetime


def copy_cell_data(source_file, target_file, source_cell, target_cell):
    """
    从源Excel文件读取指定单元格数据，写入到目标Excel文件的指定位置

    :param source_file: 源Excel文件路径
    :param target_file: 目标Excel文件路径
    :param source_cell: 源单元格坐标(例如: 'A1')
    :param target_cell: 目标单元格坐标(例如: 'B2')
    """
    try:
        # 加载源工作簿
        source_wb = load_workbook(source_file)
        source_ws = source_wb.active

        # 读取源单元格数据
        cell_value = source_ws[source_cell].value

        # 加载目标工作簿
        target_wb = load_workbook(target_file)
        target_ws = target_wb.active

        # 将数据写入目标单元格
        target_ws[target_cell] = cell_value

        # 保存目标工作簿
        target_wb.save(target_file)
        print(f"成功将数据 '{cell_value}' 从 {source_file} 的 {source_cell} 复制到 {target_file} 的 {target_cell}")

    except Exception as e:
        print(f"操作失败: {str(e)}")


# # 使用示例
# if __name__ == "__main__":
#     # 设置文件路径和单元格位置
#     source_file = "1111.xlsx"  # 源Excel文件
#     target_file = "2222.xlsx"  # 目标Excel文件
#
#     # 从源文件的A1单元格读取数据，写入到目标文件的D14单元格
#     copy_cell_data(source_file, target_file, "A1", "A1")


def duqu_excel(excel_file='1111.xlsx', excel_cell='A6', sheet_name='Sheet1'):
    """
    :param excel_file: excel文件路径
    :param excel_cell: 单元格坐标
    :param sheet_name: 工作表名称
    :return: 返回该坐标的值（前5个字符，如果是字符串）
    """
    wb = None
    try:
        # 加载 Excel 文件
        wb = load_workbook(excel_file)

        # 选择工作表
        sheet = wb[sheet_name]

        # 获取单元格值
        cell_value = sheet[excel_cell].value
        return cell_value

    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        return None

    finally:
        # 确保文件被关闭
        if wb:
            wb.close()


# 写入
def xieru(target_file, target_cell, data_value):
    """
    从源Excel文件读取指定单元格数据，写入到目标Excel文件的指定位置
    :param target_file: 目标Excel文件路径
    :param target_cell: 目标单元格坐标(例如: 'B2')
    """
    try:

        # 加载目标工作簿
        target_wb = load_workbook(target_file)
        target_ws = target_wb.active
        target_ws[target_cell] = data_value   # 将数据写入目标单元格
        target_wb.save(target_file)    # 保存目标工作簿


    except Exception as e:
        print(f"操作失败: {str(e)}")


# 清理重复打卡函数
def clean_attendance_records(time_list):
    """
    ['07:11', '11:32', '11:50', '12:07', '12:40', '12:59', '17:03']
    这是一天的考勤打卡记录，请帮我清理11：30分到13：00之间的重复打卡
    要求：
    1、保留整个list的第一次打卡和最后一次打卡
    2、只保留11：30分到13：00之间的第一次打卡和最后一次打卡
    :param time_list:
    :return:
    """
    if len(time_list) == 0:
        return []
    if len(time_list) == 1:
        return time_list

    low = '11:30'
    high = '13:00'
    first_overall = time_list[0]
    last_overall = time_list[-1]

    interval_times = [t for t in time_list if low <= t <= high]

    if not interval_times:
        return [first_overall, last_overall]

    first_interval = min(interval_times)
    last_interval = max(interval_times)

    result = [first_overall]

    if first_interval != first_overall and first_interval != last_overall:
        result.append(first_interval)

    if last_interval != last_overall and last_interval != first_interval:
        result.append(last_interval)

    result.append(last_overall)

    return result

# 读取数据

# 工号
gh_c5=duqu_excel('1111.xlsx', 'C5')
print(gh_c5)
# xieru('2222.xlsx', 'A4', gh_c5)

# 姓名
name_k5=duqu_excel('1111.xlsx', 'K5')
print(name_k5)
# xieru('2222.xlsx', 'B4', name_k5)

# 读取考勤数据
result1h = duqu_excel('1111.xlsx', 'F60')
# 将读取到的单元格数据，以5个字符分割并写入到一个list中
my_list=[result1h[i:i+5] for i in range(0, len(result1h), 5)]
print(my_list[0])   #第一次打卡
# xieru('2222.xlsx', 'E4', my_list[0])

print(my_list[-1])  #最后一次打卡
# xieru('2222.xlsx', 'H4', my_list[-1])

if len(my_list)>4:
    print('注意：有重复打卡')
    # print(my_list)
    clean_list=clean_attendance_records(my_list)
    print(clean_list)
    clean_list.insert(0, name_k5)  # 在索引0的位置插入元素1
    clean_list.insert(0, gh_c5)  # 在索引0的位置插入元素1
    print(clean_list)






# 如果打卡次数减去首次打卡和最后打卡，剩余一次时放到早上下班的位置
# if len(my_list)-2<2:
#
#     xieru('2222.xlsx', 'F4', my_list[-2])





#
#
# # 早上，上班时间
# print(result1h[0:5])
# xieru('2222.xlsx', 'E4', result1h[0:5])
# # 早上下班时间
# print(result1h[5:10])
# xieru('2222.xlsx', 'F4', result1h[5:10])
# # 中文上班
# print(result1h[10:15])
#
# t13=datetime.strptime(result1h[10:15], "%H:%M").time()
#
# if t13>datetime.strptime('17:00', "%H:%M").time():
#     xieru('2222.xlsx', 'H4', result1h[10:15])
#     print("第三组大于17，写到下班卡")
# else:
#     xieru('2222.xlsx', 'G4', result1h[10:15])
#     print("正常")
# # 中午下班
# print(result1h[15:20],'在这')
#
# # 判断中午下班卡是否没打，如果没打则跳过写入，因为上面已经写了，如果打了则写入，上卖弄就不写了
# if len(result1h[15:20]) == 0:
#     print('漏打为空')
# else:
#     xieru('2222.xlsx', 'H4', result1h[15:20])
#
# # 早上上班、下班写入
# xieru('2222.xlsx', 'D4', result[0:5])
# xieru('2222.xlsx', 'D5', result[5:10])




# # 将字符串类型转为时间对象
#
# time_obj = datetime.strptime(result[10:15], "%H:%M").time()
# if time_obj > datetime.strptime('15:00', "%H:%M").time():
#     print("中午上班卡忘记打了")
#     xieru('2222.xlsx', 'D6', '未打卡')
#     print("本次打卡为下班卡，应该放到D7")
#     xieru('2222.xlsx', 'D7', result[10:15])
#     print("写入成功")
# else:
#     xieru('2222.xlsx', 'D6', result[10:15])
