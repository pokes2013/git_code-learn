from openpyxl import load_workbook, Workbook
from itertools import dropwhile
import datetime
from typing import List, Union
from datetime import time


def process_attendance_records(records: List[Union[str, datetime.time]]) -> List[Union[str, datetime.time]]:
    """
    处理打卡记录，按照规则筛选有效打卡时间

    Args:
        records: 原始打卡记录列表，格式为 [工号, 姓名, 公司, 日期, 时间1, 时间2, ...]

    Returns:
        处理后的打卡记录列表，包含筛选后的有效打卡时间
    """
    # 提取时间部分（从索引4开始），并过滤掉None值
    times = [t for t in records[4:] if t is not None]

    # 如果没有打卡时间，直接返回原记录
    if not times:
        return records

    result_times = []  # 存储处理后的有效打卡时间

    # 1. 早上上班：取第一次打卡（如果早于9:00）
    morning_candidates = [t for t in times if t < time(9, 0)]
    if morning_candidates:
        result_times.append(min(morning_candidates))  # 取最早的一个

    # 2. 早上下班：最接近11:30且大于11:30
    target_am_end = time(10, 59)
    am_end_candidates = [t for t in times if t > target_am_end]
    if am_end_candidates:
        # 计算最接近目标时间的时间点
        am_end = min(am_end_candidates,
                     key=lambda t: abs((datetime.datetime.combine(datetime.date.today(), t) -
                                        datetime.datetime.combine(datetime.date.today(),
                                                                  target_am_end)).total_seconds()))
        # 确保时间顺序正确
        if not result_times or am_end > result_times[-1]:
            result_times.append(am_end)

    # 3. 下午上班：最接近13:00且小于13:00
    target_pm_start = time(13, 0)
    pm_start_candidates = [t for t in times if time(11, 30) < t < target_pm_start]
    if pm_start_candidates:
        pm_start = max(pm_start_candidates)  # 取最接近13:00的时间
        # 确保时间顺序正确
        if not result_times or pm_start > result_times[-1]:
            result_times.append(pm_start)

    # 4. 下午下班：最接近17:00且大于17:00
    target_pm_end = time(16, 59)
    pm_end_candidates = [t for t in times if t > target_pm_end]
    if pm_end_candidates:
        # 计算最接近目标时间的时间点
        pm_end = min(pm_end_candidates,
                     key=lambda t: abs((datetime.datetime.combine(datetime.date.today(), t) -
                                        datetime.datetime.combine(datetime.date.today(),
                                                                  target_pm_end)).total_seconds()))
        # 确保时间顺序正确
        if not result_times or pm_end > result_times[-1]:
            result_times.append(pm_end)

    # 5. 晚上加班：取最后一次打卡（在17:00之后）
    overtime_candidates = [t for t in times if t > time(16, 59)]
    if overtime_candidates:
        overtime_end = max(overtime_candidates)  # 取最晚的一个
        # 确保时间顺序正确
        if not result_times or overtime_end > result_times[-1]:
            result_times.append(overtime_end)

    # 返回前4个字段（非时间字段）加上处理后的时间字段
    return records[:4] + result_times


# 加载Excel文件
wb = load_workbook('5555.xlsx')
sheet = wb['Sheet2']

# 获取所有行数据（跳过表头）
all_rows = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]

# 处理数据
processed_data = []
for row_data in all_rows:
    # 去除末尾的None值
    cleaned_list = list(dropwhile(lambda x: x is None, reversed(row_data)))[::-1]
    # 处理打卡记录
    cleaned_list = process_attendance_records(cleaned_list)

    # 检查早上上班打卡是否缺失（11:00后打卡视为缺卡）
    if isinstance(cleaned_list[4], time) and cleaned_list[4] > time(11, 0):
        cleaned_list.insert(4, "缺卡")

    # 检查早上下班打卡是否缺失（11:30-12:30无打卡）
    if not any(time(11, 30) <= t <= time(12, 30) for t in cleaned_list if isinstance(t, time)):
        cleaned_list.insert(5, "缺卡")

    # 检查下午上班打卡是否缺失（11:40-14:00无打卡）
    if not any(time(11, 40) <= t <= time(14, 0) for t in cleaned_list if isinstance(t, time)):
        cleaned_list.insert(6, "缺卡")  # 修正索引位置

    # 处理加班打卡（保留最晚的一个，删除其他）
    overtime_indices = [(i, t) for i, t in enumerate(cleaned_list) if isinstance(t, time) and i >= 7]
    if len(overtime_indices) > 1:
        # 找出最晚打卡的索引
        latest_idx = max(overtime_indices, key=lambda x: x[1])[0]
        # 删除其他加班打卡（从后往前删除避免索引变化）
        for i, _ in sorted([idx for idx in overtime_indices if idx[0] != latest_idx], reverse=True):
            cleaned_list.pop(i)

    processed_data.append(cleaned_list)

# 创建新的工作簿并写入处理后的数据
wb_output = Workbook()
ws_output = wb_output.active

# 添加表头
headers = ['ID', '姓名', '公司', '日期', '上班时间', '上班状态', '下班状态', '下班时间']
ws_output.append(headers)

# 逐行写入处理后的数据
for row in processed_data:
    # 转换时间对象为字符串
    processed_row = [item.strftime('%H:%M:%S') if isinstance(item, time) else item for item in row]
    ws_output.append(processed_row)

# 保存Excel文件
wb_output.save('output2.xlsx')
print("数据已成功写入到 output.xlsx")