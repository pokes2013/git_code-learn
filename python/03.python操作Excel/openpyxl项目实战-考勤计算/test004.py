import openpyxl
from openpyxl import load_workbook
from datetime import datetime


def copy_cell_data(source_file, target_file, source_cell, target_cell):
    """
    从源Excel文件读取指定单元格数据，写入到目标Excel文件的指定位置

    :param source_file: 源Excel文件路径
    :param target_file: 目标Excel文件路径
    :param source_cell: 源单元格坐标(例如: 'A1')
    :param target_cell: 目标单元格坐标(例如: 'B2')
    """
    try:
        # 加载源工作簿
        source_wb = load_workbook(source_file)
        source_ws = source_wb.active

        # 读取源单元格数据
        cell_value = source_ws[source_cell].value

        # 加载目标工作簿
        target_wb = load_workbook(target_file)
        target_ws = target_wb.active

        # 将数据写入目标单元格
        target_ws[target_cell] = cell_value

        # 保存目标工作簿
        target_wb.save(target_file)
        print(f"成功将数据 '{cell_value}' 从 {source_file} 的 {source_cell} 复制到 {target_file} 的 {target_cell}")

    except Exception as e:
        print(f"操作失败: {str(e)}")


# # 使用示例
# if __name__ == "__main__":
#     # 设置文件路径和单元格位置
#     source_file = "1111.xlsx"  # 源Excel文件
#     target_file = "2222.xlsx"  # 目标Excel文件
#
#     # 从源文件的A1单元格读取数据，写入到目标文件的D14单元格
#     copy_cell_data(source_file, target_file, "A1", "A1")


def duqu_excel(excel_file='1111.xlsx', excel_cell='A6', sheet_name='Sheet1'):
    """
    :param excel_file: excel文件路径
    :param excel_cell: 单元格坐标
    :param sheet_name: 工作表名称
    :return: 返回该坐标的值（前5个字符，如果是字符串）
    """
    wb = None
    try:
        # 加载 Excel 文件
        wb = load_workbook(excel_file)

        # 选择工作表
        sheet = wb[sheet_name]

        # 获取单元格值
        cell_value = sheet[excel_cell].value
        return cell_value

    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        return None

    finally:
        # 确保文件被关闭
        if wb:
            wb.close()


# 写入
def xieru(target_file, target_cell, data_value):
    """
    从源Excel文件读取指定单元格数据，写入到目标Excel文件的指定位置
    :param target_file: 目标Excel文件路径
    :param target_cell: 目标单元格坐标(例如: 'B2')
    """
    try:

        # 加载目标工作簿
        target_wb = load_workbook(target_file)
        target_ws = target_wb.active
        target_ws[target_cell] = data_value   # 将数据写入目标单元格
        target_wb.save(target_file)    # 保存目标工作簿


    except Exception as e:
        print(f"操作失败: {str(e)}")


# 清理重复打卡函数
def clean_attendance_records(time_list):
    """
    ['07:11', '11:32', '11:50', '12:07', '12:40', '12:59', '17:03']
    这是一天的考勤打卡记录，请帮我清理11：30分到13：00之间的重复打卡
    要求：
    1、保留整个list的第一次打卡和最后一次打卡
    2、只保留11：30分到13：00之间的第一次打卡和最后一次打卡
    :param time_list:
    :return:
    """
    if len(time_list) == 0:
        return []
    if len(time_list) == 1:
        return time_list

    low = '11:30'
    high = '13:00'
    first_overall = time_list[0]
    last_overall = time_list[-1]

    interval_times = [t for t in time_list if low <= t <= high]

    if not interval_times:
        return [first_overall, last_overall]

    first_interval = min(interval_times)
    last_interval = max(interval_times)

    result = [first_overall]

    if first_interval != first_overall and first_interval != last_overall:
        result.append(first_interval)

    if last_interval != last_overall and last_interval != first_interval:
        result.append(last_interval)

    result.append(last_overall)

    return result

# 读取数据

info=[]
# 工号

get_id=duqu_excel('1111.xlsx', 'C5')
info.append(get_id)
print(get_id)

# xieru('2222.xlsx', 'A4', gh_c5)

# 姓名
get_name=duqu_excel('1111.xlsx', 'K5')
info.append(get_name)
print(get_name)
# xieru('2222.xlsx', 'B4', name_k5)

# 日期

get_date=str(duqu_excel('1111.xlsx', 'B4'))+"号"
info.append(get_date)
print(get_date)

print(info)


# 读取考勤数据
result1h = duqu_excel('1111.xlsx', 'B6')
# 将读取到的单元格数据，以5个字符分割并写入到一个list中
data_list=[result1h[i:i+5] for i in range(0, len(result1h), 5)]

# 重复打卡处理
if len(data_list)>4:
    print('注意：有重复打卡')
    # print(my_list)
    data_list=clean_attendance_records(data_list)
    print("处理重复",data_list)

# 缺卡处理
if len(data_list)<4:
    data_list.insert(2, "缺卡")  # 在索引0的位置插入元素1
    print("处理缺卡：",data_list)




print(data_list)


# data_list.insert(0, get_date)  # 在索引0的位置插入元素1
# data_list.insert(0, get_name)  # 在索引0的位置插入元素1
# data_list.insert(0, get_id)  # 在索引0的位置插入元素1

he_list=[info,data_list]
print(he_list)


# min_daka = data_list[0]  # 第一次打卡
# max_daka = data_list[-1]  # 最后一次打卡

# if len(data_list)>4:
#     print('注意：有重复打卡')
#     # print(my_list)
#     clean_list=clean_attendance_records(data_list)
#     print(clean_list)
#     clean_list.insert(0, get_date)  # 在索引0的位置插入元素1
#     clean_list.insert(0, get_name)  # 在索引0的位置插入元素1
#     clean_list.insert(0, get_id)  # 在索引0的位置插入元素1
#     print(clean_list)



