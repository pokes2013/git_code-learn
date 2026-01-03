from openpyxl import load_workbook
from itertools import dropwhile
import datetime
from typing import List, Union
from datetime import time

def process_attendance_records(records: List[Union[str, datetime.time]]) -> List[Union[str, datetime.time]]:
    """
    处理打卡记录，按照规则筛选有效打卡时间

    Args:
        records: 原始打卡记录列表，格式为 [工号, 姓名, 公司, 日期, 时间1, 时间2, ...]

    Returns:
        处理后的打卡记录列表，包含筛选后的有效打卡时间
    """
    # 提取时间部分（从索引4开始），并过滤掉None值
    times = [t for t in records[4:] if t is not None]

    if not times:
        return records

    result_times = []

    # 1. 早上上班：取第一次打卡（如果早于9:00）
    morning_candidates = [t for t in times if t < datetime.time(9, 0)]
    if morning_candidates:
        morning_start = min(morning_candidates)  # 取最早的一个
        result_times.append(morning_start)

    # 2. 早上下班：最接近11:30且大于11:30
    target_am_end = datetime.time(10, 59)
    am_end_candidates = [t for t in times if t > target_am_end]
    if am_end_candidates:
        am_end = min(am_end_candidates,
                     key=lambda t: (datetime.datetime.combine(datetime.date.today(), t) -
                                    datetime.datetime.combine(datetime.date.today(), target_am_end)).total_seconds())

        # 确保早上下班在早上上班之后（如果有早上上班）
        if not result_times or am_end > result_times[-1]:
            result_times.append(am_end)

    # 3. 下午上班：最接近13:00且小于13:00
    target_pm_start = datetime.time(13, 0)
    pm_start_candidates = [t for t in times if t < target_pm_start and t > datetime.time(11, 30)]
    if pm_start_candidates:
        pm_start = max(pm_start_candidates)  # 最接近13:00的小于13:00的时间

        # 确保下午上班在早上下班之后（如果有早上下班）
        if not result_times or pm_start > result_times[-1]:
            result_times.append(pm_start)

    # 4. 下午下班：最接近17:00且大于17:00
    target_pm_end = datetime.time(16, 59)
    pm_end_candidates = [t for t in times if t > target_pm_end]
    if pm_end_candidates:
        pm_end = min(pm_end_candidates,
                     key=lambda t: (datetime.datetime.combine(datetime.date.today(), t) -
                                    datetime.datetime.combine(datetime.date.today(), target_pm_end)).total_seconds())

        # 确保下午下班在合理的时间顺序之后
        if not result_times or pm_end > result_times[-1]:
            result_times.append(pm_end)

    # 5. 晚上加班：取最后一次打卡（在17:00之后）
    overtime_candidates = [t for t in times if t > datetime.time(16, 59)]
    if overtime_candidates:
        overtime_end = max(overtime_candidates)  # 取最晚的一个

        # 确保加班时间在下午下班之后（如果有下午下班）
        if not result_times or overtime_end > result_times[-1]:
            result_times.append(overtime_end)

    return records[:4] + result_times


# 加载 Excel 文件
wb = load_workbook('fenlie.xlsx')
sheet = wb['Sheet1']

# 一次性获取并转换为列表
# all_rows = [list(row) for row in sheet.iter_rows(min_row=2, max_row=1600, values_only=True)]
all_rows = [list(row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)]

# 处理数据
k_data=[]
for i, row_data in enumerate(all_rows, 1):
    # print(row_data)
    # print("-"*200)
    # 去除控制None
    cleaned_list = list(dropwhile(lambda x: x is None, reversed(row_data)))[::-1]
    # print(f"第{i}行数据: {cleaned_list}")
    cleaned_list=process_attendance_records(cleaned_list)

    # print(cleaned_list)

    # 处理早上上班打卡
    if isinstance(cleaned_list[4], datetime.time) and cleaned_list[4] > datetime.time(11, 0):
        # print("早上班缺卡")
        cleaned_list.insert(4, "缺卡")
        # cleaned_list.insert(4, "缺卡")
    # print(cleaned_list)
    # #
    # 处理早上下班打卡

    # 检查11:30-12:00是否有打卡，没有则插入缺卡
    if not any(time(11, 30) <= t <= time(12, 30) for t in cleaned_list if isinstance(t, time)):
        cleaned_list.insert(5, "缺卡")

    # 处理下午上班
    # 检查11:40-13:00是否有打卡，没有则插入缺卡
    if not any(time(11, 40) <= t <= time(14, 0) for t in cleaned_list if isinstance(t, time)):
        cleaned_list.insert(5, "缺卡")


    # 处理加班
    overtime = [(i, t) for i, t in enumerate(cleaned_list) if isinstance(t, datetime.time) and i >= 7]

    if len(overtime) > 1:
        # 找出最晚打卡的索引
        latest_idx = max(overtime, key=lambda x: x[1])[0]
        # 删除其他加班打卡（从后往前）
        [cleaned_list.pop(i) for i, _ in sorted(overtime, reverse=True) if i != latest_idx]

    print(cleaned_list)
    k_data.append(cleaned_list)

# 写入数据到新的文件

# 创建新的工作簿和工作表
from openpyxl import Workbook
wb2 = Workbook()
ws2 = wb2.active

# 添加表头（如果需要）
headers = ['ID', '姓名', '公司', '日期', '上班时间', '上班状态', '下班状态', '下班时间']
ws2.append(headers)

# 逐行写入数据
for row in k_data:
    # 处理datetime.time对象，转换为字符串格式
    processed_row = []
    for item in row:
        if isinstance(item, datetime.time):
            processed_row.append(item.strftime('%H:%M:%S'))
        else:
            processed_row.append(item)
    ws2.append(processed_row)

# 保存Excel文件
wb2.save('output.xlsx')
print("数据已成功写入到 output.xlsx")
