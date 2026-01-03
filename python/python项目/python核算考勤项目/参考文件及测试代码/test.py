from openpyxl import load_workbook
from itertools import dropwhile
import datetime
from typing import List, Union
from datetime import time, datetime as dt
import re


def convert_to_time(time_value) -> time:
    """
    将时间值转换为time对象

    Args:
        time_value: 可能是字符串或time对象的时间值

    Returns:
        time对象
    """
    if isinstance(time_value, time):
        return time_value
    elif isinstance(time_value, str):
        # 处理不同格式的时间字符串
        try:
            # 清理字符串，移除空格等
            time_value = time_value.strip()
            if not time_value:
                return None

            if len(time_value) <= 5:  # HH:MM格式
                return dt.strptime(time_value, "%H:%M").time()
            else:  # HH:MM:SS格式
                return dt.strptime(time_value, "%H:%M:%S").time()
        except ValueError:
            # 如果解析失败，返回None
            return None
    return None


def extract_times_from_string(time_string) -> List[str]:
    """
    从连接的时间字符串中提取独立的时间

    Args:
        time_string: 连接的时间字符串，如 '07:2911:3112:5817:12'

    Returns:
        分离后的时间字符串列表，如 ['07:29', '11:31', '12:58', '17:12']
    """
    if not isinstance(time_string, str):
        return []

    # 使用正则表达式匹配时间格式 HH:MM
    # 查找所有形如 HH:MM 的模式
    time_pattern = r'[0-2]?\d:[0-5]\d'
    times = re.findall(time_pattern, time_string)

    return times


def process_attendance_records(records: List[Union[str, time]]) -> List[Union[str, time]]:
    """
    处理打卡记录，按照规则筛选有效打卡时间

    Args:
        records: 原始打卡记录列表，格式为 [工号, 姓名, 公司, 日期, 时间字符串]

    Returns:
        处理后的打卡记录列表，包含筛选后的有效打卡时间
    """
    # 如果记录不足5个元素，直接返回
    if len(records) < 5:
        return records

    # 提取所有时间字符串并分离
    all_time_strings = []
    for cell_value in records[4:]:  # 从索引4开始处理时间相关字段
        if isinstance(cell_value, str):
            # 从连接的时间字符串中提取独立时间
            extracted_times = extract_times_from_string(cell_value)
            all_time_strings.extend(extracted_times)
        elif isinstance(cell_value, time):
            # 如果已经是time对象，转换回字符串
            all_time_strings.append(cell_value.strftime('%H:%M'))

    # 转换所有时间为time对象
    times = []
    for time_str in all_time_strings:
        converted_time = convert_to_time(time_str)
        if converted_time is not None:
            times.append(converted_time)

    if not times:
        return records[:4]  # 只返回基本信息

    # 按时间排序
    times.sort()

    result_times = []

    # 1. 早上上班：取第一次打卡（如果早于9:00）
    morning_candidates = [t for t in times if t < time(9, 0)]
    if morning_candidates:
        morning_start = min(morning_candidates)  # 取最早的一个
        result_times.append(morning_start)

    # 2. 早上下班：最接近10:59且大于10:59
    target_am_end = time(10, 59)
    am_end_candidates = [t for t in times if t > target_am_end]
    if am_end_candidates:
        am_end = min(am_end_candidates,
                     key=lambda t: (dt.combine(dt.today(), t) -
                                    dt.combine(dt.today(), target_am_end)).total_seconds())

        # 确保早上下班在早上上班之后（如果有早上上班）
        if not result_times or am_end > result_times[-1]:
            result_times.append(am_end)

    # 3. 下午上班：11:30到13:00之间最接近13:00的时间
    target_pm_start = time(13, 0)
    pm_start_candidates = [t for t in times if t < target_pm_start and t > time(11, 30)]
    if pm_start_candidates:
        pm_start = max(pm_start_candidates)  # 最接近13:00的小于13:00的时间

        # 确保下午上班在早上下班之后（如果有早上下班）
        if not result_times or pm_start > result_times[-1]:
            result_times.append(pm_start)

    # 4. 下午下班：最接近16:59且大于16:59
    target_pm_end = time(16, 59)
    pm_end_candidates = [t for t in times if t > target_pm_end]
    if pm_end_candidates:
        pm_end = min(pm_end_candidates,
                     key=lambda t: (dt.combine(dt.today(), t) -
                                    dt.combine(dt.today(), target_pm_end)).total_seconds())

        # 确保下午下班在合理的时间顺序之后
        if not result_times or pm_end > result_times[-1]:
            result_times.append(pm_end)

    # 5. 晚上加班：取最后一次打卡（在16:59之后）
    overtime_candidates = [t for t in times if t > time(16, 59)]
    if overtime_candidates:
        overtime_end = max(overtime_candidates)  # 取最晚的一个

        # 确保加班时间在下午下班之后（如果有下午下班）
        if not result_times or overtime_end > result_times[-1]:
            result_times.append(overtime_end)

    return records[:4] + result_times


# 加载 Excel 文件
wb = load_workbook('melted_cities.xlsx')
sheet = wb['Sheet1']

# 一次性获取并转换为列表
all_rows = [list(row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)]

# 处理数据
k_data = []
for i, row_data in enumerate(all_rows, 1):
    # 去除控制None
    cleaned_list = list(dropwhile(lambda x: x is None, reversed(row_data)))[::-1]
    processed_list = process_attendance_records(cleaned_list)

    # 处理早上上班打卡
    if len(processed_list) > 4 and isinstance(processed_list[4], time) and processed_list[4] > time(11, 0):
        processed_list.insert(4, "缺卡")

    # 处理早上下班打卡
    # 检查11:30-12:30是否有打卡，没有则插入缺卡
    if len(processed_list) > 4:
        time_values = [t for t in processed_list[4:] if isinstance(t, time)]
        if not any(time(11, 30) <= t <= time(12, 30) for t in time_values):
            # 确保不会在索引越界的情况下插入
            insert_pos = min(5, len(processed_list))
            processed_list.insert(insert_pos, "缺卡")

    # 处理下午上班
    # 检查11:40-14:00是否有打卡，没有则插入缺卡
    if len(processed_list) > 5:
        time_values = [t for t in processed_list[4:] if isinstance(t, time)]
        if not any(time(11, 40) <= t <= time(14, 0) for t in time_values):
            # 确保不会在索引越界的情况下插入
            insert_pos = min(6, len(processed_list))
            processed_list.insert(insert_pos, "缺卡")

    # 处理加班
    overtime = [(idx, t) for idx, t in enumerate(processed_list) if isinstance(t, time) and idx >= 4]

    if len(overtime) > 1:
        # 找出最晚打卡的索引
        latest_idx = max(overtime, key=lambda x: x[1])[0]
        # 删除其他加班打卡（从后往前删除，避免索引变化）
        for idx, _ in sorted(overtime, reverse=True):
            if idx != latest_idx and idx < len(processed_list):
                processed_list.pop(idx)

    print(processed_list)
    k_data.append(processed_list)

# 写入数据到新的文件

# 创建新的工作簿和工作表
from openpyxl import Workbook

wb2 = Workbook()
ws2 = wb2.active

# 添加表头（如果需要）
headers = ['ID', '姓名', '公司', '日期', '上班时间', '上班状态', '下班状态', '下班时间']
ws2.append(headers)

# 逐行写入数据
for row in k_data:
    # 处理datetime.time对象，转换为字符串格式
    processed_row = []
    for item in row:
        if isinstance(item, time):
            processed_row.append(item.strftime('%H:%M:%S'))
        else:
            processed_row.append(item)
    ws2.append(processed_row)

# 保存Excel文件
wb2.save('chaifenlie.xlsx')
print("数据已成功写入到 chaifenlie.xlsx")
