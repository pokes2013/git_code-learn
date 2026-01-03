# 如果整天每打卡就会在表格中看不见，需要修复这个BUG


from openpyxl import load_workbook
from datetime import time, datetime
from openpyxl import Workbook


# 给1-9号的日期前面补0
def format_date(date_str):
    """
    格式化日期字符串，给1-9号的日期前面补0

    参数:
    date_str: 日期字符串，如 "3号", "24号"等

    返回:
    str: 格式化后的日期字符串
    """
    if not date_str or not isinstance(date_str, str):
        return date_str

    if date_str.endswith('号'):
        # 提取数字部分
        num_part = date_str[:-1]
        if num_part.isdigit():
            num = int(num_part)
            # 给1-9号的日期前面补0
            if 1 <= num <= 9:
                return f"{num:02d}号"

    return date_str


# 获取最接近目标时间的两次不同打卡记录
def get_closest_punches_advanced(row, target_time1_str="11:30:00", target_time2_str="13:00:00", start_index=2):
    """
    高级版本：获取最接近目标时间的两次不同打卡记录，考虑多种情况

    返回:
    tuple: (最接近11:30的时间, 最接近13:00的时间, 说明信息)
    """
    if len(row) <= start_index:
        return None, None, "无打卡记录"

    # 过滤有效的打卡时间（排除包含"缺卡"的文字）
    valid_times = []
    for time_str in row[start_index:]:
        if (time_str and isinstance(time_str, str) and
                ':' in time_str and '缺卡' not in time_str and
                len(time_str.split(':')) == 3):
            try:
                time_obj = datetime.strptime(time_str, '%H:%M:%S')
                valid_times.append((time_str, time_obj))
            except ValueError:
                continue

    if not valid_times:
        return None, None, "无有效打卡记录"

    if len(valid_times) == 1:
        return valid_times[0][0], None, "只有一次有效打卡"

    # 转换目标时间
    target_time1 = datetime.strptime(target_time1_str, '%H:%M:%S')
    target_time2 = datetime.strptime(target_time2_str, '%H:%M:%S')

    # 找到最接近11:30的时间
    closest_to_11_30 = min(valid_times, key=lambda x: abs(x[1] - target_time1))

    # 从剩余时间中找到最接近13:00的时间
    remaining_times = [t for t in valid_times if t[0] != closest_to_11_30[0]]
    closest_to_13_00 = min(remaining_times, key=lambda x: abs(x[1] - target_time2))

    return closest_to_11_30[0], closest_to_13_00[0]


# 保留指定时间范围内的第一次打卡
def merge_checkins_in_range(attendance_row, start_range="06:00:00", end_range="08:00:00"):
    """
    处理员工考勤记录，合并指定时间范围内的重复打卡记录，只保留最早的一次

    参数:
    attendance_row: 列表，包含员工一天的考勤记录，格式如:
                   ['姓名', '日期', '打卡时间1', '打卡时间2', ...]
    start_range: 字符串，时间范围起始点，格式为"HH:MM:SS"，默认为"06:00:00"
    end_range: 字符串，时间范围结束点，格式为"HH:MM:SS"，默认为"08:00:00"

    返回:
    列表，合并后的考勤记录，指定时间范围内的重复打卡只保留最早一次
    """
    # 分离基本信息和打卡时间
    name = attendance_row[0]
    date = attendance_row[1]
    check_times = attendance_row[2:]

    # 筛选出指定时间范围内的打卡时间并找到最早的一次
    range_checkins = [t for t in check_times if start_range <= t < end_range]
    earliest_in_range = min(range_checkins) if range_checkins else None

    # 筛选出指定时间范围外的打卡时间
    other_checkins = [t for t in check_times if not (start_range <= t < end_range)]

    # 构建结果列表
    result = [name, date]

    # 如果有指定时间范围内的打卡，添加最早的一次
    if earliest_in_range:
        result.append(earliest_in_range)

    # 添加其他时间的打卡记录
    result.extend(other_checkins)

    return result


# 保留指定时间范围内的最后一次打卡
def keep_last_evening_punch(row, start_time="16:00:00", end_time="23:00:00"):
    """
    保留指定时间范围内的最后一次打卡

    参数:
    row: 数据行
    start_time: 开始时间，默认"16:00:00"
    end_time: 结束时间，默认"23:00:00"

    返回:
    list: 处理后的列表
    """
    if len(row) <= 2:
        return row

    base_info = row[:2]
    all_times = row[2:]

    evening_times = []
    other_times = []

    for time in all_times:
        if time and start_time <= time <= end_time:
            evening_times.append(time)
        else:
            other_times.append(time)

    last_evening = evening_times[-1] if evening_times else None

    if last_evening:
        return base_info + other_times + [last_evening]
    else:
        return base_info + other_times


# 获取时间范围内的打卡次数
def count_evening_punches_compact(row, start_time, end_time):
    """
    简洁版本：使用列表推导式,获取时间范围内的打卡次数
    """
    if len(row) <= 2:
        return 0

    evening_punches = [time for time in row[2:] if time and start_time <= time <= end_time]
    return len(evening_punches)


data = []
# 加载 Excel 文件
wb = load_workbook('../003-SplitDataTimetype.xlsx')
ws = wb.active  # 获取活动工作表
# 遍历每一行（从第二行开始，假设第一行是标题）
xinlist = []
for row in ws.iter_rows(min_row=1, values_only=True):
    row = list(row)
    row = [row[1]] + row[3:]  # 选择第2列和第5列及之后的所有列
    # 过滤掉 None 值
    row = [item for item in row if item is not None]

    # 处理日期
    row[1] = format_date(row[1])
    # print(row)

    # 一、对于缺卡的处理
    # 01.上班卡
    row = merge_checkins_in_range(row)   # 将字符串转换为时间对象
    count = count_evening_punches_compact(row, start_time="04:00:00", end_time="11:30:00")  # 早上上班如果大于10点打卡

    if count == 0:
        row.insert(2, "早上请假")

    # 02.早上下班和下午上班的处理
    # 闲的时候优化一下，可以使用11.30分的下标定位，判断：下标+1，下标+2 来确定后面的打卡到底是重复还是下午上班卡。

    count = count_evening_punches_compact(row, start_time="11:00:00", end_time="13:50:00")

    if count < 2:
        if count == 0:
            row.insert(3, "早上下班缺卡")
            row.insert(4, "下午上班缺卡")
        else:
            row.insert(4, "下午上班缺卡")
    elif count > 2:
        for _ in range(count - 2):
            if len(row) > 4: del row[4]

    # 03.最后一次打卡的处理
    # 获取17-23.59之间的打卡次数
    zuihouchongfu_list = [time for time in row[5:] if time and '17:00:00' <= time <= '23:59:00']
    punch_count2 = len(zuihouchongfu_list)

    count = count_evening_punches_compact(row, start_time="16:00:00", end_time="23:59:59")

    if count > 1:
        row = keep_last_evening_punch(row)
    if count == 0:
        row.insert(5, "下午下班缺卡")

    # print(row)
    data.append(row)

print(data)

# 写入数据到新的文件
# 创建一个工作簿
wb = Workbook()
# 获取活动工作表
ws = wb.active
# 可以给工作表命名
ws.title = "考勤记录"

# 添加表头（如果需要的话）
ws.append(['姓名', '日期', '早上上班', '早上下班', '下午上班', '下午打卡&最后一次打卡'])

# 遍历数据并添加到工作表
for row in data:
    ws.append(row)

# 保存文件
wb.save('004-CelldataMove.xlsx')
print("数据已成功写入：004-CelldataMove.xlsx")