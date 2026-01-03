from openpyxl import load_workbook
from datetime import time, datetime
from typing import List, Union, Any
import pandas as pd


def process_attendance_records(records: List[Union[str, datetime.time]]) -> List[Union[str, datetime.time]]:
    """
    处理打卡记录，按照规则筛选有效打卡时间

    规则说明：
    - 早上上班：取早于9:00的最早打卡时间
    - 早上下班：取最接近11:30且大于11:30的打卡时间
    - 下午上班：取最接近13:00且小于13:00的打卡时间
    - 下午下班：取最接近17:00且大于17:00的打卡时间
    - 晚上加班：取晚于17:00的最晚打卡时间

    Args:
        records: 原始打卡记录列表，格式为 [工号, 姓名, 公司, 日期, 时间1, 时间2, ...]
                前4个元素为员工基本信息，从第5个元素开始为打卡时间

    Returns:
        处理后的打卡记录列表，包含筛选后的有效打卡时间
        格式为 [工号, 姓名, 公司, 日期, 有效时间1, 有效时间2, ...]
    """
    # 提取时间部分并去重，保持原始顺序
    times = []
    seen = set()
    for t in records[4:]:
        if t is not None and t not in seen:
            times.append(t)
            seen.add(t)

    # 如果没有有效打卡时间，直接返回原始记录
    if not times:
        return records

    # 对时间进行排序，便于后续处理
    times_sorted = sorted(times)
    result_times = []

    # 1. 处理早上上班打卡：取第一次打卡（如果早于9:00）
    morning_candidates = [t for t in times_sorted if t < datetime.time(9, 0)]
    if morning_candidates:
        result_times.append(morning_candidates[0])  # 最早的一个

    # 2. 处理早上下班打卡：最接近11:30且大于11:30
    target_am_end = datetime.time(11, 30)
    am_end_candidates = [t for t in times_sorted if t > target_am_end]
    if am_end_candidates:
        # 找到第一个大于11:30的时间（已排序，所以第一个就是最接近的）
        am_end = am_end_candidates[0]
        if not result_times or am_end > result_times[-1]:
            result_times.append(am_end)

    # 3. 处理下午上班打卡：最接近13:00且小于13:00
    target_pm_start = datetime.time(13, 0)
    # 筛选所有早于13:00且晚于11:30的打卡时间
    pm_start_candidates = [t for t in times_sorted
                           if datetime.time(11, 30) < t < target_pm_start]
    if pm_start_candidates:
        pm_start = pm_start_candidates[-1]  # 最大的一个（最接近13:00）
        if not result_times or pm_start > result_times[-1]:
            result_times.append(pm_start)

    # 4. 处理下午下班打卡：最接近17:00且大于17:00
    target_pm_end = datetime.time(17, 0)
    pm_end_candidates = [t for t in times_sorted if t > target_pm_end]
    if pm_end_candidates:
        pm_end = pm_end_candidates[0]  # 第一个大于17:00的时间（最接近）
        if not result_times or pm_end > result_times[-1]:
            result_times.append(pm_end)

    # 5. 处理晚上加班打卡：取最后一次打卡（在17:00之后）
    overtime_candidates = [t for t in times_sorted if t > datetime.time(17, 0)]
    if overtime_candidates:
        overtime_end = overtime_candidates[-1]  # 最晚的一个
        if not result_times or overtime_end > result_times[-1]:
            result_times.append(overtime_end)

    return records[:4] + result_times


def process_row_data(row_data: List[Union[str, int, None]]) -> List[Any]:
    """
    处理行数据：去除None值，并将时间字符串转换为时间对象

    Args:
        row_data: 包含可能为None值的原始数据列表

    Returns:
        处理后的列表，None值被移除，时间字符串转换为datetime.time对象
    """
    # 去除None值
    filtered_data = [item for item in row_data if item is not None]

    # 处理时间字符串（从索引4开始，即第5个元素）
    result = filtered_data[:4]  # 保留前4个元素不变

    # 转换剩余元素中的时间字符串
    for item in filtered_data[4:]:
        if isinstance(item, str) and ':' in item and len(item.split(':')) == 3:
            try:
                # 尝试将字符串转换为时间对象
                time_obj = datetime.strptime(item, '%H:%M:%S').time()
                result.append(time_obj)
            except ValueError:
                # 如果转换失败，保留原字符串
                result.append(item)
        else:
            # 非时间字符串，直接保留
            result.append(item)

    return result


# 获取并转换为列表
wb = load_workbook('../003-SplitDataTimetype.xlsx')
sheet = wb['Sheet1']
all_rows = [list(row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)]

# 处理数据异常：检索缺卡、重复打卡
k_data = []
for i, row_data in enumerate(all_rows, 1):
    # 处理重复打卡
    cleaned_list = process_row_data(row_data)

    # 一、处理缺卡
    # 01、处理早上上班打卡
    if cleaned_list[4] > time(11, 00):
        cleaned_list.insert(4, "缺卡")

    # 02.处理早上下班打卡，检查11:30-12:00是否有打卡，没有则插入缺卡
    if not any(time(11, 30) <= t <= time(11, 50) for t in cleaned_list if isinstance(t, time)):
        cleaned_list.insert(5, "缺卡")

    # 03.处理下午上班，检查11:40-13:00是否有打卡，没有则插入缺卡
    if not any(time(11, 40) < t <= time(13, 30) for t in cleaned_list if isinstance(t, time)):
        cleaned_list.insert(6, "缺卡")

    # 二、处理加班
    overtime = [(i, t) for i, t in enumerate(cleaned_list) if isinstance(t, time) and i >= 7]
    print(cleaned_list)

    if len(overtime) > 1:
        # 找出最晚打卡的索引
        latest_idx = max(overtime, key=lambda x: x[1])[0]
        # 删除其他加班打卡（从后往前）
        [cleaned_list.pop(i) for i, _ in sorted(overtime, reverse=True) if i != latest_idx]
    k_data.append(cleaned_list)

# 写入数据到新的文件

# 添加表头（如果需要）
headers = ['ID', '姓名', '公司', '日期', '上班时间', '上班状态', '下班状态', '下班时间']
k_data.insert(0, headers)

# 写入文件
df = pd.DataFrame(k_data)
df.to_excel('004-CelldataMove.xlsx', index=False, header=False)
print('写入完成！')

# 打印结果查看
for row in k_data:
    print(row)
