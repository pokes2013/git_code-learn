from openpyxl import load_workbook
from datetime import datetime

# 加载 Excel 文件
wb = load_workbook('1111.xlsx')

# 获取所有工作表名称
print(wb.sheetnames)  # ['Sheet1', 'Sheet2', ...]

# 选择工作表
sheet = wb['Sheet1']  # 通过名称选择
# 或者
sheet = wb.active  # 获取活动工作表

# 方法1：通过单元格坐标
# 标体
cell_value_biaoti = sheet['A1'].value
print(cell_value_biaoti)
# 考勤时间范围
cell_value_A3 = sheet['A3'].value
cell_value_C3 = sheet['C3'].value
print(cell_value_A3, cell_value_C3)

# 读取工号
cell_value_C5 = sheet['C5'].value
print(cell_value_C5)
cell_value_K5 = sheet['K5'].value
print(cell_value_K5)

# # 读取整行
# row_data = []
# for cell in sheet[4]:  # 读取第一行
#     row_data.append(cell.value)
# print(row_data)

# 读取其中一天的打卡数据,每月2号
cell_value_B6 = sheet['B6'].value
# print(cell_value_B6)
# 单元格长度，标准20
if len(cell_value_B6) < 20:
    print("警告：打开次数异常,缺卡", len(cell_value_B6))

print(cell_value_B6[0:5])
print(cell_value_B6[5:10])
print(cell_value_B6[10:15])
print(cell_value_B6[15:20])

# 将字符串转为时间类型
time_obj = datetime.strptime(cell_value_B6[0:5], "%H:%M").time()

# 判断实际打开时间
if time_obj > datetime.strptime("07:30", "%H:%M").time():
    print("警告：早上上班打开异常，初次判断为迟到")
    if time_obj > datetime.strptime("11:30", "%H:%M").time():
        print("二次判断：上午请假")
else:
    pass

print("没迟到")

# if cell_value_B6[0:5]>7:30:
