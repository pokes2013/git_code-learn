from openpyxl import load_workbook, Workbook
from datetime import datetime


# 格式化日期，给1-9号补0
def format_date(date_str):
    if isinstance(date_str, str) and date_str.endswith('号'):  # 检查是否为有效日期字符串
        num = date_str[:-1]
        if num.isdigit() and 1 <= int(num) <= 9:  # 1-9号补0处理
            return f"{int(num):02d}号"
    return date_str  # 无需处理的日期直接返回


# 合并指定时间范围内的打卡记录，保留最早一次
def merge_morning_checkins(attendance_row, start="06:00:00", end="08:00:00"):
    name, date = attendance_row[:2]  # 提取姓名和日期
    check_times = attendance_row[2:]  # 提取所有打卡时间
    in_range = [t for t in check_times if start <= t < end]  # 筛选指定时段打卡
    others = [t for t in check_times if not (start <= t < end)]  # 其他时段打卡
    result = [name, date]  # 初始化结果列表
    if in_range: result.append(min(in_range))  # 添加最早打卡(如有)
    result.extend(others)  # 添加其他时段打卡
    return result


# 保留指定时间范围内的最后一次打卡
def keep_last_evening_checkin(row, start="16:00:00", end="23:00:00"):
    base_info = row[:2]  # 基本信息(姓名、日期)
    all_times = row[2:]  # 所有打卡时间
    evening = [t for t in all_times if t and start <= t <= end]  # 筛选晚间打卡
    others = [t for t in all_times if not (t and start <= t <= end)]  # 其他时间打卡
    return base_info + others + ([evening[-1]] if evening else [])  # 组合结果


# 统计指定时间范围内的打卡次数
def count_checkins_in_range(row, start, end):
    return len([t for t in row[2:] if t and start <= t <= end]) if len(row) > 2 else 0


# 主程序
data = []  # 存储处理后的考勤数据
wb = load_workbook('003-SplitDataTimetype.xlsx')  # 加载源Excel文件
ws = wb.active  # 获取活动工作表

# 遍历数据行(从第2行开始，跳过标题行)
for row in ws.iter_rows(min_row=1, values_only=True):
    row = list(row)  # 转换为列表便于处理
    row = [row[1]] + row[3:]  # 提取第2列和第5列及之后的数据
    row = [item for item in row if item is not None]  # 过滤空值

    if len(row) >= 2: row[1] = format_date(row[1])  # 格式化日期(如存在)

    # 处理早上上班打卡
    row = merge_morning_checkins(row)
    # 如无早上打卡记录，标记"早上请假"
    if count_checkins_in_range(row, "04:00:00", "11:30:00") == 0:
        row.insert(2, "早上请假")

    # 处理午间打卡(早上下班和下午上班)
    noon_count = count_checkins_in_range(row, "11:00:00", "13:50:00")
    if noon_count < 2:  # 打卡次数不足
        if noon_count == 0:  # 无午间打卡
            row.insert(3, "早上下班缺卡")
            row.insert(4, "下午上班缺卡")
        else:  # 只有一次午间打卡
            row.insert(4, "下午上班缺卡")
    elif noon_count > 2:  # 打卡次数过多，保留前2次
        for _ in range(noon_count - 2):
            if len(row) > 4: del row[4]

    # 处理下午下班打卡
    evening_count = count_checkins_in_range(row, "16:00:00", "23:59:59")
    if evening_count > 1:  # 多次打卡保留最后一次
        row = keep_last_evening_checkin(row)
    if evening_count == 0:  # 无下午打卡记录
        row.insert(5, "下午下班缺卡")

    data.append(row)  # 添加到结果列表

# 创建新Excel文件并写入数据
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "考勤记录"  # 工作表命名
# 添加表头
ws_new.append(['姓名', '日期', '早上上班', '早上下班', '下午上班', '最后一次打卡'])
# 写入处理后的数据
for row in data: ws_new.append(row)
# 保存文件
wb_new.save('004-CelldataMove.xlsx')
print("数据已成功写入：004-CelldataMove.xlsx")
