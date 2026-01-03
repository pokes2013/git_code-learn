from openpyxl import load_workbook
from datetime import datetime
import pandas as pd


def split_time_string_to_datetime(s):
    # 每5个字符分割一次并转换为datetime.time对象
    time_list = []
    for i in range(0, len(s), 5):
        time_str = s[i:i + 5]
        # 将字符串转换为datetime对象
        time_obj = datetime.strptime(time_str, '%H:%M').time()
        time_list.append(time_obj)
    return time_list


# 加载 Excel 文件
wb = load_workbook(r'002-NTSdata.xlsx')
sheet = wb['Sheet1']

# 一次性获取并转换为列表
all_rows = [list(row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)]

chaifen_list = []

for row in all_rows:
    if row[4] is not None:
        kaoqin_time_list = split_time_string_to_datetime(row[4]) if row[4] is not None else []

        # 直接创建平铺的列表，而不是嵌套列表
        data = row[0:4] + kaoqin_time_list

        # 只有当有考勤时间数据时才添加到结果列表
        if len(data) > 4:
            chaifen_list.append(data)

# 写入文件
df = pd.DataFrame(chaifen_list)
df.to_excel('003-SplitDataTimetype.xlsx', index=False, header=False)
print('写入完成！')


# 打印结果查看
for row in chaifen_list:
    print(row)