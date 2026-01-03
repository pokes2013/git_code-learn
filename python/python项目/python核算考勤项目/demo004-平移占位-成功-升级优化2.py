from openpyxl import load_workbook
from datetime import time, datetime
from openpyxl import Workbook


# 格式化日期，给1-9号补0
def format_date(date_str):
    if isinstance(date_str, str) and date_str.endswith('号'):  # 检查是否为有效日期字符串
        num = date_str[:-1]
        if num.isdigit() and 1 <= int(num) <= 9:  # 1-9号补0处理
            return f"{int(num):02d}号"
    return date_str  # 无需处理的日期直接返回


# 自定义排序key函数
def sort_key(row):
    name = row[0]  # 人名在第一列
    # 提取日期数字并转换为整数
    if len(row) > 1 and isinstance(row[1], str) and '号' in row[1]:
        day = int(row[1].replace('号', ''))
    else:
        day = 0  # 默认值
    return (name, day)


# 生成完整的日期范围（假设一个月有31天）
def generate_all_days():
    return [f"{i:02d}号" for i in range(1, 32)]


# 统计指定时间范围内的打卡次数
def count_checkins_in_range(row, start, end):
    return len([t for t in row[2:] if t and start <= t <= end]) if len(row) > 2 else 0

# 合并指定时间范围内的打卡记录，保留最早一次
def merge_morning_checkins(attendance_row, start="06:00:00", end="08:00:00"):
    name, date = attendance_row[:2]  # 提取姓名和日期
    check_times = attendance_row[2:]  # 提取所有打卡时间
    in_range = [t for t in check_times if start <= t < end]  # 筛选指定时段打卡
    others = [t for t in check_times if not (start <= t < end)]  # 其他时段打卡
    result = [name, date]  # 初始化结果列表
    if in_range: result.append(min(in_range))  # 添加最早打卡(如有)
    result.extend(others)  # 添加其他时段打卡
    return result

#判断时间在某个范围内
def is_time_in_range(check_time, start_time, end_time):
    """
    判断时间是否在指定范围内

    参数:
    check_time (datetime): 需要判断的时间
    start_time (datetime): 范围开始时间
    end_time (datetime): 范围结束时间

    返回:
    str: "yes" 如果在范围内，否则返回"NO"
    """
    # 检查输入是否有效
    if not all([check_time, start_time, end_time]):
        return "NO"  # 输入无效时返回NO

    # 处理跨天的情况（如晚上10点到早上6点）
    if end_time < start_time:
        # 时间在start_time到午夜 或 午夜到end_time 都算在范围内
        return "yes" if check_time >= start_time or check_time <= end_time else "NO"
    else:
        # 正常情况，时间在start_time和end_time之间
        return "yes" if start_time <= check_time <= end_time else "NO"

# 主程序
data = []  # 存储处理后的考勤数据
wb = load_workbook('003-SplitDataTimetype.xlsx')  # 加载源Excel文件
ws = wb.active  # 获取活动工作表

# 收集所有人名和对应的数据
person_data = {}  # 结构: {人名: {日期: 数据行}}

# 首先获取所有人名
all_names = set()
for row in ws.iter_rows(min_row=1, values_only=True):  # 从第2行开始，跳过标题
    if row[1]:  # 确保第2列（人名）不为空
        all_names.add(row[1])

# 初始化每个人的数据结构
all_days = generate_all_days()
for name in all_names:
    person_data[name] = {}

# 遍历数据行处理每个人
for row in ws.iter_rows(min_row=1, values_only=True):  # 从第2行开始，跳过标题
    # if row[1] == '唐有华':
        if row[1]:  # 确保有姓名
            name = row[1]
            row = list(row)  # 转换为列表便于处理
            row = [row[1]] + row[3:]  # 提取第2列和第5列及之后的数据
            row = [item for item in row if item is not None]  # 过滤空值


            if len(row) >= 2:
                row[1] = format_date(row[1])  # 格式化日期(如存在)

            # 按日期存储数据
            if len(row) > 1:
                person_data[name][row[1]] = row

        # 01.处理早上上班卡
        zaoshang_count = count_checkins_in_range(row, "04:00:00", "10:00:00")   #打卡次数
        zuizaoyici_rs=merge_morning_checkins(row, "04:00:00", "11:00:00")  #范围内取最早一次打卡


        if zaoshang_count>2:   #大于0，说明打了早上卡，但是还有重复的打卡
            del row[3]
        if zaoshang_count>1:   #大于0，说明打了早上卡，但是还有重复的打卡
            del row[3]

        if zaoshang_count < 1:
            # print('早上上班缺卡')
            row.insert(2, '早上班缺卡')

        # 02.处理早上下班卡
        zhchifan_count = count_checkins_in_range(row, "10:40:00", "15:00:00")  # 打卡次数
        if zhchifan_count > 0:
            if zhchifan_count == 1 :
                # if is_time_in_range(row[2],"04:00:00", "09:00:00") == 'yes' :   #判断11点到15点打卡次数为1，并且判断早上是否有打卡
                row.insert(3, '早下班缺卡')
                # if count_checkins_in_range(row, "16:20:00", "23:59:00")>1:
                #      row.insert(4, '午上班缺卡')

            if zhchifan_count == 3:
                del row[4]
            if zhchifan_count == 4:
                del row[4]
                del row[4]

        elif zhchifan_count == 0:
            row.insert(3, '早下班缺卡')
            row.insert(4, '午上班缺卡')

        # 03.处理下午上班
        # wushangban_count = count_checkins_in_range(row, "11:20:00", "15:00:00")  # 打卡次数
        # if wushangban_count < 1:
        #     row.insert(4, '午上班缺卡')


        # 04.处理下班卡
        xiaban_count = count_checkins_in_range(row, "16:00:00", "23:59:00")

        if xiaban_count>0:   # 有下班卡
            if xiaban_count == 3:   #下班卡重复
                del row[-2]         #删除倒数第二个
                del row[-2]
            if xiaban_count == 2:  # 下班卡重复
                del row[-2]
        else:               #无下班卡
            # print('没打下班卡')
            row.append('午下班缺卡')

        # if xiaban_count > 1:
        #     del row[-2]
        # print(row)

# 为每个人生成完整的数据集，包含占位行
for name in all_names:
    for day in all_days:
        if day in person_data[name]:
            data.append(person_data[name][day])
        else:
            # 插入占位行：[人名, 日期, '无数据']
            placeholder_row = [name, day, '休息']
            data.append(placeholder_row)


# 对收集到的数据进行排序
sorted_data = sorted(data, key=sort_key)

# 打印排序结果
print("排序后的数据（包含所有人的占位行）：")
for row in sorted_data:
    print(row)

# 如果需要保存到新的Excel文件
wb_new = Workbook()
ws_new = wb_new.active

# 写入标题行（可选）
ws_new.append(['姓名', '日期', '早上上班','早上下班',  '下午上班', '最后打卡'])

for i, row in enumerate(sorted_data, 2):  # 从第2行开始写入数据
    for j, value in enumerate(row, 1):
        ws_new.cell(row=i, column=j, value=value)

    # 如果是有数据的行，继续写入时间信息
    if len(row) > 2 and row[2] != '无数据':
        for k in range(3, len(row)):
            ws_new.cell(row=i, column=k + 1, value=row[k])

wb_new.save('004-CelldataMove.xlsx')
print(f"数据已保存到 004-CelldataMove.xlsx")
print(f"处理了 {len(all_names)} 个人：{list(all_names)}")